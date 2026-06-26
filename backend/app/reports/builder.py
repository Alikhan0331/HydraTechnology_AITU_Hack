"""Report builders: CSV, Excel (xlsx) and PDF over a list of structures.

All three respect whatever filtered list is passed in, so the same filters used
in the UI produce the matching export ("export current view").
"""
import csv
import io
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from ..enums import CONDITIONS, RISK_LABELS, ConditionCode, RiskLevel

FONT_PATH = Path(__file__).parent / "fonts" / "DejaVuSans.ttf"

COND_LABEL = {c.value: CONDITIONS[c][0] for c in ConditionCode}
COND_COLOR = {c.value: CONDITIONS[c][1] for c in ConditionCode}
RISK_LABEL = {r.value: RISK_LABELS[r] for r in RiskLevel}

# (header, attribute, getter)
COLUMNS = [
    ("ID", "id"),
    ("Наименование", "name"),
    ("Тип", "type"),
    ("Район", "district"),
    ("Состояние", lambda s: COND_LABEL.get(s.condition, s.condition)),
    ("Риск", lambda s: RISK_LABEL.get(s.risk_level, s.risk_level)),
    ("Балл риска", "risk_score"),
    ("Год", "year_built"),
    ("Длина, км", "length_km"),
    ("Износ, %", "wear_percent"),
    ("Водоисточник", "water_source"),
    ("Посл. осмотр", "last_inspection"),
    ("След. осмотр", "next_inspection"),
    ("Широта", "latitude"),
    ("Долгота", "longitude"),
]


def _val(s, accessor):
    v = accessor(s) if callable(accessor) else getattr(s, accessor, None)
    return "" if v is None else v


# --------------------------------------------------------------------------- #
#  CSV
# --------------------------------------------------------------------------- #
def build_csv(structures) -> bytes:
    buf = io.StringIO()
    buf.write("﻿")  # BOM so Excel reads UTF-8 Cyrillic correctly
    w = csv.writer(buf, delimiter=";")
    w.writerow([h for h, _ in COLUMNS])
    for s in structures:
        w.writerow([_val(s, acc) for _h, acc in COLUMNS])
    return buf.getvalue().encode("utf-8")


# --------------------------------------------------------------------------- #
#  Excel
# --------------------------------------------------------------------------- #
def build_xlsx(structures, summary: dict | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Объекты"

    header_fill = PatternFill("solid", fgColor="1D4ED8")
    header_font = Font(color="FFFFFF", bold=True)
    for col, (h, _acc) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

    for r, s in enumerate(structures, start=2):
        for col, (_h, acc) in enumerate(COLUMNS, start=1):
            v = _val(s, acc)
            ws.cell(row=r, column=col, value=str(v) if isinstance(v, date) else v)

    widths = [6, 30, 16, 16, 18, 12, 10, 8, 10, 9, 16, 13, 13, 10, 10]
    for i, wd in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = wd
    ws.freeze_panes = "A2"

    if summary:
        ws2 = wb.create_sheet("Сводка")
        ws2["A1"] = "Показатель"; ws2["B1"] = "Значение"
        ws2["A1"].font = ws2["B1"].font = Font(bold=True)
        rows = [
            ("Всего объектов", summary.get("total")),
            ("Индекс состояния (0-100)", summary.get("overall_condition_index")),
            ("Общая протяжённость, км", summary.get("total_length_km")),
            ("Средний возраст, лет", summary.get("avg_age_years")),
        ]
        for code in ConditionCode:
            rows.append((COND_LABEL[code.value], summary.get("by_condition", {}).get(code.value, 0)))
        for i, (k, v) in enumerate(rows, start=2):
            ws2.cell(row=i, column=1, value=k)
            ws2.cell(row=i, column=2, value=v)
        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 16

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# --------------------------------------------------------------------------- #
#  PDF
# --------------------------------------------------------------------------- #
def build_pdf(structures, summary: dict | None = None, title: str | None = None) -> bytes:
    from fpdf import FPDF

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_font("DejaVu", "", str(FONT_PATH))
    pdf.add_font("DejaVu", "B", str(FONT_PATH))
    pdf.add_page()

    pdf.set_font("DejaVu", "B", 16)
    pdf.cell(0, 9, title or "Отчёт по гидротехническим сооружениям", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("DejaVu", "", 9)
    pdf.set_text_color(110, 110, 110)
    pdf.cell(0, 6, f"Жамбылская область · сформировано {date.today().isoformat()} · объектов: {len(structures)}",
             new_x="LMARGIN", new_y="NEXT")
    pdf.set_text_color(0, 0, 0)
    pdf.ln(2)

    if summary:
        pdf.set_font("DejaVu", "B", 11)
        pdf.cell(0, 7, "Сводка", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("DejaVu", "", 9)
        kpi = (
            f"Всего: {summary.get('total')}   |   "
            f"Индекс состояния: {summary.get('overall_condition_index')}/100   |   "
            f"Протяжённость: {summary.get('total_length_km')} км   |   "
            f"Средний возраст: {summary.get('avg_age_years')} лет"
        )
        pdf.cell(0, 6, kpi, new_x="LMARGIN", new_y="NEXT")
        bc = summary.get("by_condition", {})
        cond_line = "   ".join(f"{COND_LABEL[c.value]}: {bc.get(c.value, 0)}" for c in ConditionCode)
        pdf.cell(0, 6, cond_line, new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

    # table (cap rows so the PDF stays reasonable; sorted by caller)
    cols = [("Наименование", 70), ("Тип", 34), ("Район", 34),
            ("Состояние", 38), ("Риск", 24), ("Год", 16), ("Длина", 18)]
    pdf.set_font("DejaVu", "B", 9)
    pdf.set_fill_color(29, 78, 216)
    pdf.set_text_color(255, 255, 255)
    for h, w in cols:
        pdf.cell(w, 8, h, border=0, fill=True, align="L")
    pdf.ln()
    pdf.set_text_color(20, 20, 20)
    pdf.set_font("DejaVu", "", 8)

    limit = 70
    for i, s in enumerate(structures[:limit]):
        fill = i % 2 == 1
        pdf.set_fill_color(244, 247, 250)
        row = [
            (str(s.name)[:42], 70), (str(s.type)[:18], 34), (str(s.district)[:18], 34),
            (COND_LABEL.get(s.condition, s.condition), 38),
            (RISK_LABEL.get(s.risk_level, s.risk_level), 24),
            (str(s.year_built or "—"), 16),
            (f"{s.length_km:.1f}" if s.length_km else "—", 18),
        ]
        for text, w in row:
            pdf.cell(w, 7, text, border=0, fill=fill, align="L")
        pdf.ln()

    if len(structures) > limit:
        pdf.ln(2)
        pdf.set_text_color(110, 110, 110)
        pdf.cell(0, 6, f"… и ещё {len(structures) - limit} объектов (полный список — в Excel/CSV).",
                 new_x="LMARGIN", new_y="NEXT")

    return bytes(pdf.output())


def build_object_passport(obj, risk_eval: dict, inspections, repairs) -> bytes:
    """One-page PDF passport for a single object (object card download)."""
    from fpdf import FPDF

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=14)
    pdf.add_font("DejaVu", "", str(FONT_PATH))
    pdf.add_font("DejaVu", "B", str(FONT_PATH))
    pdf.add_page()

    pdf.set_font("DejaVu", "B", 16)
    pdf.cell(0, 9, "Паспорт гидротехнического сооружения", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("DejaVu", "B", 13)
    pdf.set_text_color(29, 78, 216)
    pdf.cell(0, 8, str(obj.name), new_x="LMARGIN", new_y="NEXT")
    pdf.set_text_color(0, 0, 0)
    pdf.ln(2)

    def row(label, value):
        pdf.set_font("DejaVu", "B", 10)
        pdf.cell(55, 7, label)
        pdf.set_font("DejaVu", "", 10)
        pdf.cell(0, 7, str(value), new_x="LMARGIN", new_y="NEXT")

    row("Тип", obj.type)
    row("Район", obj.district)
    row("Состояние", COND_LABEL.get(obj.condition, obj.condition))
    row("Risk Score", f"{risk_eval['risk_score']} / 100 — {risk_eval['risk_level']}")
    row("Год ввода", obj.year_built or "—")
    row("Длина, км", obj.length_km if obj.length_km is not None else "—")
    row("Износ, %", obj.wear_percent if obj.wear_percent is not None else "—")
    row("Водоисточник", obj.water_source or "—")
    row("Координаты", f"{obj.latitude}, {obj.longitude}")
    row("Посл. осмотр", obj.last_inspection or "—")
    row("След. осмотр", obj.next_inspection or "—")
    pdf.ln(2)

    pdf.set_font("DejaVu", "B", 11)
    pdf.cell(0, 7, "Причины риска", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("DejaVu", "", 9)
    for reason in risk_eval["risk_reasons"]:
        pdf.cell(0, 6, f"•  {reason}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    for title, items, fmt in (
        ("История осмотров", inspections,
         lambda x: f"{x.date} — {x.inspection_type}: {(x.notes or '')[:68]}"),
        ("История ремонтов", repairs,
         lambda x: f"{x.repair_date} — {x.repair_type}: {(x.notes or '')[:68]}"),
    ):
        pdf.set_font("DejaVu", "B", 11)
        pdf.cell(0, 7, title, new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("DejaVu", "", 9)
        if not items:
            pdf.cell(0, 6, "—", new_x="LMARGIN", new_y="NEXT")
        for it in items[:12]:
            pdf.cell(0, 6, fmt(it), new_x="LMARGIN", new_y="NEXT")
        pdf.ln(1)

    pdf.ln(2)
    pdf.set_font("DejaVu", "", 8)
    pdf.set_text_color(120, 120, 120)
    pdf.cell(0, 5, f"Сформировано {date.today().isoformat()} · HydraTechnology",
             new_x="LMARGIN", new_y="NEXT")
    return bytes(pdf.output())
