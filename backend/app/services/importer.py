"""Bulk import of structures from CSV / Excel (TZ task 2: integrate data from
multiple sources).

Built to swallow "dataset-like" files (e.g. the hackathon датасет.xls): leading
title rows are skipped, the header row is auto-detected, Russian column names are
matched by substring, "удов./не удов." is turned into a 4-level condition, and
rows without coordinates are geocoded to a real Zhambyl district. Every row is
deduplicated against the existing catalog before insertion (TZ task 4).
"""
import csv
import io
import re

from sqlalchemy import select
from sqlalchemy.orm import Session

from .. import crud, models, schemas
from ..enums import CONDITIONS, DISTRICT_NAMES, DISTRICTS, ConditionCode
from .classification import derive_condition
from .discovery import _name_similar
from .geo import coords_for_district, haversine_m

# canonical field -> key substrings to look for inside a header cell.
# Order matters: the first column that matches a field wins (so "Год ввода"
# binds to year_built before a later "Год принятия на баланс" can).
_SYNONYMS = {
    "name": ["наименование", "название", "name", "имя", "объект"],
    "type": ["тип", "вид сооружения", "type"],
    "latitude": ["широта", "lat"],
    "longitude": ["долгота", "lon", "lng"],
    "year_built": ["год ввода", "год постройки", "year", "год"],
    "length_km": ["протяж", "длина", "length"],
    "condition": ["техническое состояние", "состояние", "condition"],
    "wear_percent": ["износ", "wear"],
    "water_source": ["водоисточник", "источник", "река", "water"],
    "capacity": ["пропускная", "расход", "capacity"],
    "district": ["район", "расположение", "district"],
}

_COND_IN = {c.value: c.value for c in ConditionCode}
_COND_IN.update({CONDITIONS[c][0].lower(): c.value for c in ConditionCode})
_COND_IN.update({"исправно": "good", "норма": "good", "критическое": "emergency"})

DUP_SPOT_M = 350        # same coordinates → same object
DUP_NAME_M = 8000       # similar name + same type within this radius → duplicate


def _norm(s) -> str:
    return " ".join(str(s).lower().split()) if s is not None else ""


def _match_field(header: str) -> str | None:
    h = _norm(header)
    if not h:
        return None
    for field, keys in _SYNONYMS.items():
        if any(k in h for k in keys):
            return field
    return None


def _num(v):
    if v is None:
        return None
    s = str(v).strip().replace(",", ".").replace("%", "").replace(" ", "")
    if not s or s.lower() in ("nan", "none", "-"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _int(v):
    f = _num(v)
    return int(f) if f is not None else None


# --- generic table parsing (header detection) -------------------------------
def _rows_from_csv(content: bytes) -> list[list]:
    text = content.decode("utf-8-sig", errors="replace")
    delim = ";" if text[:2048].count(";") >= text[:2048].count(",") else ","
    return [list(r) for r in csv.reader(io.StringIO(text), delimiter=delim)]


def _rows_from_xlsx(content: bytes) -> list[list]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    return [list(r) for r in wb.active.iter_rows(values_only=True)]


def _rows_from_xls(content: bytes) -> list[list]:
    """Legacy .xls (e.g. the original датасет.xls).
    Auto-detects xlsx disguised as .xls by checking ZIP magic bytes (PK header).
    """
    # xlsx files are ZIP archives and start with PK\x03\x04
    if content[:4] == b'PK\x03\x04':
        return _rows_from_xlsx(content)
    import xlrd

    book = xlrd.open_workbook(file_contents=content)
    sh = book.sheet_by_index(0)
    return [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]


def _detect_header(rows: list[list]) -> int:
    best_i, best_score = 0, 0
    for i, row in enumerate(rows[:15]):
        score = sum(1 for c in row if _match_field(c))
        if score > best_score:
            best_i, best_score = i, score
    return best_i if best_score >= 2 else 0


def parse_file(filename: str, content: bytes) -> list[dict]:
    """Return a list of {canonical_field: value} dicts, one per data row."""
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        rows = _rows_from_xlsx(content)
    elif name.endswith(".xls"):
        rows = _rows_from_xls(content)
    else:
        rows = _rows_from_csv(content)
    if not rows:
        return []
    hi = _detect_header(rows)
    header = rows[hi]
    col_field: dict[int, str] = {}
    for idx, cell in enumerate(header):
        field = _match_field(cell)
        if field and field not in col_field.values():   # first column per field wins
            col_field[idx] = field

    out = []
    for row in rows[hi + 1:]:
        rec = {}
        for idx, field in col_field.items():
            if idx < len(row) and row[idx] not in (None, ""):
                rec[field] = row[idx]
        if rec:
            rec["_col0"] = row[0] if row and row[0] not in (None, "") else None
            out.append(rec)
    return out


# --- normalization ----------------------------------------------------------
def _condition_from(raw, year_built, wear_fraction) -> str | None:
    if raw is None:
        return None
    n = _norm(raw)
    if n in _COND_IN:
        return _COND_IN[n]
    if n.startswith("не удов") or "неудов" in n:
        return derive_condition("unsatisfactory", year_built, wear_fraction, None, None)
    if n.startswith("удов"):
        return derive_condition("satisfactory", year_built, wear_fraction, None, None)
    return None


_TYPE_HINTS = [
    ("плотин", "Плотина"), ("дамб", "Дамба"), ("шлюз", "Шлюз"),
    ("насос", "Насосная станция"), ("водозабор", "Водозабор"),
    ("гидропост", "Гидропост"), ("канал", "Канал"),
]


def _infer_type(name: str) -> str | None:
    n = _norm(name)
    for key, ru in _TYPE_HINTS:
        if key in n:
            return ru
    return None


def _map_district(value, row_index: int) -> str:
    """Real district as-is; anonymized 'Район N'/unknown → deterministic real one."""
    if value:
        s = str(value).strip()
        if s in DISTRICTS:
            return s
        m = re.search(r"\d+", s)
        seed = int(m.group()) if m else abs(hash(s))
        return DISTRICT_NAMES[seed % len(DISTRICT_NAMES)]
    return DISTRICT_NAMES[row_index % len(DISTRICT_NAMES)]


# --- dedup against the catalog (TZ task 4 on ingest) ------------------------
def _find_duplicate(db: Session, name: str, type_code: str, lat: float, lon: float,
                    real_coords: bool):
    # 1) same coordinates → same object — ONLY when the file gave real coordinates
    #    (geocoded district-center coords would collide by chance and false-match).
    if real_coords:
        dlat = dlon = DUP_SPOT_M / 100_000
        spot = select(models.HydraulicStructure).where(
            models.HydraulicStructure.latitude.between(lat - dlat, lat + dlat),
            models.HydraulicStructure.longitude.between(lon - dlon, lon + dlon),
        )
        for s in db.scalars(spot):
            if haversine_m(lat, lon, s.latitude, s.longitude) <= DUP_SPOT_M:
                return s
    # 2) same type + matching name (primary signal for registry data without coords)
    same_type = select(models.HydraulicStructure).where(
        models.HydraulicStructure.type_code == type_code
    )
    for s in db.scalars(same_type):
        if _name_similar(name, s.name):
            return s
    return None


# --- orchestration ----------------------------------------------------------
def import_rows(db: Session, rows: list[dict], default_type: str = "Канал") -> dict:
    created, duplicates, errors = [], [], []
    skipped_empty = 0

    for i, rec in enumerate(rows, start=1):
        try:
            # skip leftover sub-header / column-number rows (condition is numeric there)
            cond_raw = rec.get("condition")
            if cond_raw is not None and _norm(cond_raw).replace(".", "").isdigit():
                skipped_empty += 1
                continue

            wear = _num(rec.get("wear_percent"))
            wear_fraction = (wear / 100) if wear is not None and wear > 1 else wear
            year = _int(rec.get("year_built"))

            name = str(rec.get("name")).strip() if rec.get("name") else None
            # explicit type column → use it; else infer from the name; else default
            type_value = (str(rec.get("type")).strip() if rec.get("type")
                          else (_infer_type(name or "") or default_type))

            lat, lon = _num(rec.get("latitude")), _num(rec.get("longitude"))
            real_coords = lat is not None and lon is not None
            has_payload = any(_num(rec.get(k)) is not None or rec.get(k)
                              for k in ("year_built", "length_km", "wear_percent",
                                        "water_source", "capacity"))
            # name from the source number (column №) so re-importing the same
            # registry is idempotent; fall back to row index.
            if not name or name.isdigit():
                if not (has_payload or real_coords):
                    skipped_empty += 1   # title / sub-header / empty row
                    continue
                src_num = _int(rec.get("_col0"))
                name = f"{type_value} №{src_num if src_num is not None else i}"

            district = _map_district(rec.get("district"), i)
            geocoded = not real_coords
            if not real_coords:
                lat, lon = coords_for_district(district, i)

            _ru, code = crud._resolve_type(type_value)
            dup = _find_duplicate(db, name, code, lat, lon, real_coords=real_coords)
            if dup:
                duplicates.append({"row": i, "name": name,
                                   "matched_id": dup.id, "matched_name": dup.name})
                continue

            payload = schemas.StructureCreate(
                name=name, type=type_value, district=district,
                latitude=lat, longitude=lon,
                condition=_condition_from(rec.get("condition"), year, wear_fraction),
                length_km=_num(rec.get("length_km")), year_built=year,
                wear_percent=wear, capacity=_num(rec.get("capacity")),
                water_source=str(rec["water_source"]).strip() if rec.get("water_source") else None,
            )
            obj = crud.create_structure(db, payload, source="imported",
                                        verification_status="needs_check")
            created.append({"row": i, "id": obj.id, "name": obj.name, "type": obj.type,
                            "condition": obj.condition, "risk_level": obj.risk_level,
                            "geocoded": geocoded})
        except Exception as exc:
            errors.append({"row": i, "reason": str(exc)[:140]})

    return {
        "total_rows": len(rows),
        "created": len(created),
        "duplicates": len(duplicates),
        "errors": len(errors),
        "skipped_empty": skipped_empty,
        "created_items": created[:100],
        "duplicate_items": duplicates[:100],
        "error_items": errors[:50],
    }


TEMPLATE_CSV = (
    "name;type;district;latitude;longitude;year_built;length_km;condition;wear_percent;water_source\n"
    "Канал Демонстрационный;Канал;Жамбылский;42.90;71.40;1985;12.5;удов.;25;р. Талас\n"
    "Плотина Пример;Плотина;Меркенский;42.87;73.18;1979;;не удов.;65;р. Мерке\n"
)
